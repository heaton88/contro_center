# -*- coding: utf-8 -*-
# _author_: heaton
# _date_: 2021/7/28 2:07 下午
# _FileName_: customer_contract.py

import requests
import json
import openpyxl
from config_control import login,access_header


wb = openpyxl.Workbook()
filename = '接口报错.xlsx'
ws1 = wb.active
ws1.title = "customer interface"
ws1['A1'] = '客户名称'
ws1['B1'] = '客户id'
ws1['C1'] = '年度产值'

customer_url = 'http://k8sdev.golowo.com/g3-enterprisecenter-web/customer/searchCustomerList'
payload = {"size": 100000000}

id_list = []
code_list = []
name_list = []
customer_year = []
customer_month = []

customer_list = []

customer_rep = requests.post(customer_url,data=json.dumps(payload),headers=access_header('json'))
customer_json = json.loads(customer_rep.content)['data']['items']

# print(customer_json)
for i in range(0,len(customer_json)):
    id_list.append(customer_json[i]['id'])
    code_list.append(customer_json[i]['code'])
    name_list.append(customer_json[i]['name'])

    # print(id_list[i],code_list[i],name_list[i])

for j in range(0,len(id_list)):
    customer_value_url = 'http://k8sdev.golowo.com/g3-screen-web/manageCenter/queryOutputValue' \
                     'Amount?orgCode=-6944734322880132047,0100000000&customerCode='+\
                         id_list[j]+','+id_list[j]
    s = requests.session()
    s.keep_alive = False
    customer_re = requests.get(customer_value_url,headers=access_header('json'),timeout=None)
    if customer_re.status_code != 200:
        print(name_list[j],"接口报错")

    else:
        for l in range(2,len(id_list)):
            ws1.cell(row=l,column=1).value = name_list[j]
            ws1.cell(row=l,column=2).value = code_list[j]
            ws1.cell(row=l,column=3).value = json.loads(customer_re.content)['data']['yearAmount']
        print(name_list[j],json.loads(customer_re.content)['data']['yearAmount'])
        customer_year.append(json.loads(customer_re.content)['data']['yearAmount'])
        customer_month.append(json.loads(customer_re.content)['data']['monthAmount'])


wb.save(filename)

